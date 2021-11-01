from openpyxl import load_workbook
import os
import argparse
from ics import Calendar, Event
from datetime import datetime, timedelta, tzinfo

ACADEMIC_PERIOD_COL = 'A'
SECTION_COL = 'B'
MEETING_TIME_COL = 'C'
LOCATION_COL = 'D'
CAPACITY_COL = 'E'
START_DATE = 'Y'
END_DATE = 'Z'

DAYS = [ 'M', 'T', 'W', 'R', 'F']
#constants
START_ROW = 3

# read the file path from the passed in --file argument
parser = argparse.ArgumentParser(description='Generate a calendar from lewins excel file because workday sux')
parser.add_argument('--file', help='The path to the excel file with the course data')
args = parser.parse_args()

# load the excel workbook
wb = load_workbook(args.file)
sheet = wb.active

# Read all the events from the workbook
events = []

c = Calendar()

for row in sheet.iter_rows(min_row=START_ROW, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    event = {}
    event["col"] = row[0].row

    for cell in row:
        if ACADEMIC_PERIOD_COL in cell.coordinate:
            event['academic_period'] = cell.value
        elif SECTION_COL in cell.coordinate:
            event['section'] = cell.value
        elif MEETING_TIME_COL in cell.coordinate:
            event['meeting_time'] = cell.value
        elif LOCATION_COL in cell.coordinate:
            event['location'] = cell.value
        elif CAPACITY_COL in cell.coordinate:
            event['capacity'] = cell.value
        elif START_DATE in cell.coordinate:
            event['start_date'] = cell.value
        elif END_DATE in cell.coordinate:
            event['end_date'] = cell.value

    events.append(event)

# parse the events to days of the week and times they happen
for event in events:
    event['class_time'] = {}
    # when does the class happen
    if event['meeting_time']:
        meeting_time = event['meeting_time'].strip().split('\n')
        event['meeting_time'] = [t.strip() for t in meeting_time]

        for meeting_time in event['meeting_time']:
            if meeting_time:
                days, event_time = meeting_time.split('|')
                
                days = days.strip().split('-')
                start_time,end_time = event_time.strip().split('-')

                for day in days:
                    event['class_time'][DAYS.index(day)+1] = {
                        'start_time': start_time.lower().strip(),
                        'end_time': end_time.lower().strip()
                    }
        

delta = timedelta(days=1)
# convert the events to ics events
for event in events:
    date = event['start_date']
    while date <= event['end_date']:
        if date.isoweekday() in event['class_time'].keys():
            e = Event()
            e.name = f"{event['section']} ROW:{event['col']}"
            e.begin = datetime.combine(date, datetime.strptime(event['class_time'][date.isoweekday()]['start_time'], '%I:%M %p').time())
            e.end = datetime.combine(date, datetime.strptime(event['class_time'][date.isoweekday()]['end_time'], '%I:%M %p').time())
            c.events.add(e)
        date += delta



with open('events.ics', 'w') as f:

    for line in str(c).splitlines():
        if line.startswith('DT'):
            line = line.replace('Z', '')
        f.write(line + '\n')